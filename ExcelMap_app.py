import tkinter as tk
from tkinter import filedialog, ttk, messagebox
from openpyxl import load_workbook, Workbook
import xml.etree.ElementTree as ET
import json


class ExcelMapping:
    def __init__(self, root):
        self.root = root
        self.excel_file = None
        self.xml_root = None
        self.worksheets = {}
        self.xml_elements = []
        self.mapping_file = "mapping.json"
        self.element_mapping = {}  # Initialize element_mapping attribute

        self.root.title("XML to Excel Mapper")

        # Create a frame for better organization
        self.frame = ttk.Frame(root, padding="20")
        self.frame.pack()

        # Select Excel File
        self.label_excel = ttk.Label(self.frame, text="Step 1: Select an Excel file:")
        self.label_excel.grid(row=0, column=0, sticky="w", padx=5, pady=5)

        self.file_button = ttk.Button(self.frame, text="Browse", command=self.select_excel_file)
        self.file_button.grid(row=0, column=1, sticky="w", padx=5, pady=5)

        # Adding separator 
        ttk.Separator(self.frame, orient=tk.HORIZONTAL).grid(row=1, columnspan=2, sticky="ew", pady=10)

        # Select XML File
        self.label_xml = ttk.Label(self.frame, text="Step 2: Select an XML file:")
        self.label_xml.grid(row=2, column=0, sticky="w", padx=5, pady=5)

        self.xml_button = ttk.Button(self.frame, text="Browse", command=self.load_xml_file)
        self.xml_button.grid(row=2, column=1, sticky="w", padx=5, pady=5)

        # Add separator
        ttk.Separator(self.frame, orient=tk.HORIZONTAL).grid(row=3, columnspan=2, sticky="ew", pady=10)

        # Step 3: Mapping Button
        # Also has command create mapping interface 
        self.mapping_button = ttk.Button(self.frame, text="Step 3: Map XML Elements to Excel", command=self.create_mapping_interface)
        self.mapping_button.grid(row=4, columnspan=2, padx=5, pady=5)

    def select_excel_file(self):
        """Function to select an Excel file. Here I have only enabled XLSM as we will be using that."""
        file_path = filedialog.askopenfilename(filetypes=[("Excel Macro-Enabled Workbook", "*.xlsm"), ("All files", "*.*")])
        if file_path:
            self.excel_file = file_path
            self.load_worksheets()

    # This will load and display all worksheet in file
    def load_worksheets(self):
        """Function to load and display all worksheets in the selected Excel file."""
        wb = load_workbook(self.excel_file)
        self.worksheets = {ws.title: ws for ws in wb.worksheets}

    def load_xml_file(self):
        """Function to load an XML file."""
        file_path = filedialog.askopenfilename(filetypes=[("XML files", "*.xml")])

        if file_path:
            self.xml_root = parse_xml(file_path)
            self.xml_elements = self.get_xml_elements(self.xml_root)

    # Function when clicking map XML to Excel
    def create_mapping_interface(self):
        """Function to create the mapping interface."""
        if not self.excel_file:
            messagebox.showwarning("Warning", "Please select an Excel file.")
            return
        if not self.xml_root:
            messagebox.showwarning("Warning", "Please select an XML file.")
            return

        self.mapping_window = tk.Toplevel(self.root)
        self.mapping_window.title("XML to Excel Mapping")

        # Load mappings if available
        self.load_mappings()

        for index, element in enumerate(self.xml_elements):
            frame = ttk.Frame(self.mapping_window, padding="10")
            frame.grid(row=index, columnspan=2, pady=5, sticky="w")

            label = ttk.Label(frame, text=f"Map {element} to:")
            label.grid(row=0, column=0, padx=5, sticky="w")

            worksheet_var = tk.StringVar()
            column_var = tk.StringVar(value="A")
            row_var = tk.IntVar(value=1)

            # Set default values based on loaded mappings
            if element in self.element_mapping:
                worksheet_var.set(self.element_mapping[element][0].get())
                column_var.set(self.element_mapping[element][1].get())
                row_var.set(self.element_mapping[element][2].get())

            worksheet_dropdown = ttk.Combobox(frame, textvariable=worksheet_var, values=list(self.worksheets.keys()), state="readonly", width=20)
            worksheet_dropdown.grid(row=0, column=1, padx=5)

            column_dropdown = ttk.Combobox(frame, textvariable=column_var, values=self.get_excel_columns(), state="readonly", width=5)
            column_dropdown.grid(row=0, column=2, padx=5)

            row_entry = ttk.Entry(frame, textvariable=row_var, width=5)
            row_entry.grid(row=0, column=3, padx=5)

            self.element_mapping[element] = (worksheet_var, column_var, row_var)

        save_button = ttk.Button(self.mapping_window, text="Save Mapping", command=self.save_mapping)
        save_button.grid(row=len(self.xml_elements), columnspan=2, pady=10)

    def save_mapping(self):
        """Function to save the mapping."""
        mappings = {}
        for element, (worksheet_var, column_var, row_var) in self.element_mapping.items():
            worksheet = worksheet_var.get()
            column = column_var.get()
            row = row_var.get()
            mappings.setdefault(worksheet, {})[element] = (column, row)

        with open(self.mapping_file, "w") as f:
            json.dump(mappings, f)

        self.generate_excel(mappings)

    def generate_excel(self, mappings):
        """Function to generate the Excel file based on mappings."""
        try:
            wb = load_workbook(self.excel_file, keep_vba=True)
        except FileNotFoundError:
            wb = Workbook(keep_vba=True)

        for worksheet_name, element_mappings in mappings.items():
            ws = wb[worksheet_name]
            for element, (column, row) in element_mappings.items():
                values = self.get_values_for_element(self.xml_root, element)
                for value in values:
                    ws[f"{column}{row}"] = value
                    row += 1

        # Save the file with the .xlsm extension
        wb.save(self.excel_file.replace(".xlsx", ".xlsm"))
        messagebox.showinfo("Success", "Excel file generated successfully.")

    def load_mappings(self):
        """Function to load the saved mappings."""
        try:
            with open(self.mapping_file, "r") as f:
                mappings = json.load(f)
            for worksheet, element_mappings in mappings.items():
                for element, (column, row) in element_mappings.items():
                    self.element_mapping[element] = (tk.StringVar(value=worksheet), tk.StringVar(value=column), tk.IntVar(value=row))
        except FileNotFoundError:
            print("No Saved File for mapping ")
        except Exception as e:
            print("Error loading mappings:", e)

    def get_values_for_element(self, xml_element, target_element):
        """Function to get values for a specific XML element."""
        values = []
        if xml_element.tag == target_element and xml_element.text:
            values.append(xml_element.text)
        for child in xml_element:
            values.extend(self.get_values_for_element(child, target_element))
        return values

    def get_xml_elements(self, xml_root):
        """Function to get XML elements from the root."""
        elements_with_values = set()

        # Recursive function to traverse XML tree and collect elements with values
        def traverse_xml_elements(element):
            if element.text and element.text.strip():  # Check if element has non-empty text
                elements_with_values.add(element.tag)
            for child in element:
                traverse_xml_elements(child)

        traverse_xml_elements(xml_root)
        return list(elements_with_values)

    def get_excel_columns(self):
        """Function to get Excel column names."""
        return ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']


def parse_xml(file_path):
    """Function to parse an XML file and return the root element."""
    tree = ET.parse(file_path)
    root = tree.getroot()
    return root


def main():
    """Main function to initialize the application."""
    root = tk.Tk()
    app = ExcelMapping(root)
    root.mainloop()


if __name__ == "__main__":
    main()
